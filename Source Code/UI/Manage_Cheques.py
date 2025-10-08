import sys
from pathlib import Path

if getattr(sys, 'frozen', False):  # Running as a PyInstaller EXE
    root_path = Path(sys.executable).parent
else:  # Running as a Python script
    root_path = Path(__file__).parent.parent
sys.path.append(str(root_path))
 
from utilities.path_utils import *
import tkinter as tk
import sqlite3
import os
from tkinter import Canvas, Entry, ttk, Button, PhotoImage, messagebox, Menu, NORMAL, END, filedialog
from tkinter.ttk import Combobox
from tkcalendar import DateEntry
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from UI.ChequePrint import PrintOnCheque

def Manage_Cheques_ui(parent = None):

    ASSETS_PATH = Path(generate_path(root_path, "UI", "assets", "frame8"))
    
    def relative_to_assets(path: str) -> Path:
        return ASSETS_PATH / Path(path)
    
    def on_close(*args):
        window.grab_release()
        window.destroy()


    window = tk.Toplevel(parent)
    window.focus()
    window.iconbitmap(generate_path("UI", "assets", "BillMates.ico"))
    window.transient(parent)
    window.grab_set()
    window.title("Manage Cheques")
    center_window(window, 1280, 720)
    window.configure(bg = "#E7EBFF")

    start_date = None
    end_date = None


    def on_date_change(event):
        quick_date.set("Custom Range")

    def get_date_range():

        dates = [row[0] for row in fetch_data("SELECT DISTINCT Cheque_Date FROM Cheques", db='Cheque.db')]

        parsed_dates = [datetime.strptime(date, '%d/%m/%Y') for date in dates]

        min_date = min(parsed_dates)
        max_date = max(parsed_dates)

        # Format the results back to dd/mm/yyyy
        min_date_str = min_date.strftime('%d/%m/%Y')
        max_date_str = max_date.strftime('%d/%m/%Y')

        return min_date_str, max_date_str

    def update_dates(event):
        if start_date is None or end_date is None:
            return
        
        selected_option = quick_date.get()
        today = datetime.now()

        Min_date, max_date = get_date_range()

        if selected_option == "All":
            start_date.set_date(Min_date)
            end_date.set_date(max_date)
            start_date.configure(state="DISABLED")
            end_date.configure(state="DISABLED")

        elif selected_option == "Today":
            start_date.set_date(today)
            end_date.set_date(today)
            
        elif selected_option == "Last 7 Days":
            start_date.set_date(today - timedelta(days=7))
            end_date.set_date(today)
            
        elif selected_option == "Last 30 Days":
            start_date.set_date(today - timedelta(days=30))
            end_date.set_date(today)
            
        elif selected_option == "This Year":
            start_date.set_date(today.replace(month=1, day=1))
            end_date.set_date(today)
            
        elif selected_option == "Last Year":
            start_date.set_date(today.replace(year=today.year - 1, month=1, day=1))
            end_date.set_date(today.replace(year=today.year - 1, month=12, day=31))
            
        elif selected_option == "Custom Range":
            # Enable the entry fields for manual date selection
            start_date.config(state=NORMAL)
            end_date.config(state=NORMAL)
            return   
         
    def update_search_field(event):
        global search_widget
        selected_option = search_at.get()

        # Destroy the previous widget (Entry or Combobox)
        for widget in window.winfo_children():
            if widget.winfo_name() == 'search_widget':
                widget.destroy()

        canvas.itemconfig(entry_bg_5, state='hidden')

        if selected_option == "Cheque Book":
            chequeBook_Name = [name[0] for name in fetch_data("SELECT book_name FROM ChequeBook", db='Cheque.db')]
            search_widget = ttk.Combobox(
                window,
                values=chequeBook_Name,
                state="readonly", 
                font=("VarelaRound Regular", 10)
            )
            search_widget.place(
                x=704.0,
                y=56.0,
                width=170.0,
                height=25.0
            )
            search_widget.set("Select")

        elif selected_option == "Cheque Type":
            # Create a Combobox for Payment Types
            payment_types = ["Select", "Account Pay", "Cross Pay"]
            search_widget = ttk.Combobox(
                window,
                values=payment_types,
                state="readonly",  # Read-only mode to prevent manual input
                font=("VarelaRound Regular", 10)
            )
            search_widget.place(
                x=704.0,
                y=56.0,
                width=170.0,
                height=25.0
            )
            search_widget.set("Select")
        
        else:
            search_widget = Entry(
                window,
                bd=0,
                bg="#FFFFFF",
                fg="#000716",
                highlightthickness=0,
                font=("VarelaRound Regular", 10)
            )
            search_widget.place(
                x=704.0,
                y=56.0,
                width=170.0,
                height=25.0
            )

    def reset_data():
        global search_widget

        start_date.set_date(datetime.now().strftime("%d/%m/%Y"))
        end_date.set_date(datetime.now().strftime("%d/%m/%Y"))
        quick_date.set("Select")
        search_at.set("Select")
        search_widget.destroy()
        search_widget = Entry(
            window,
            bd=0,
            bg="#FFFFFF",
            fg="#000716",
            highlightthickness=0,
            font=("VarelaRound Regular", 10)
        )
        search_widget.place(
            x=704.0,
            y=56.0,
            width=170.0,
            height=25.0
        )
        payment_type.set("Select")

        canvas.itemconfig(entry_bg_5, state='normal')
        
        for row in tree.get_children():
                    tree.delete(row)

    def filter_data():
        global search_widget
        global treeview_data
        f_start_date = start_date.get()
        f_end_date = end_date.get()
        f_search_at = search_at.get()
        f_search_value = search_widget.get()
        f_payment_type = payment_type.get()

        # Initialize filters and parameters
        filters = []
        cheque_params = []

        if f_start_date != "" and f_end_date != "":
            start_d_obj = datetime.strptime(f_start_date, "%d/%m/%Y")
            end_d_obj = datetime.strptime(f_end_date, "%d/%m/%Y")

            date_list = []
            current_date = start_d_obj
            while current_date <= end_d_obj:
                date_list.append(current_date.strftime("%d/%m/%Y"))  # Match the format in the database
                current_date += timedelta(days=1)

            filters.append("Cheque_Date IN (" + ",".join(["?"] * len(date_list)) + ")")

            cheque_params.extend(date_list)

        if not f_search_at == "Select" and f_search_at != "":
            
            if f_search_value == "Select" or not f_search_value:
                pass

            elif f_search_at == "Cheque Book":
                CB_id = fetch_data(f"SELECT id FROM ChequeBook WHERE book_name = '{f_search_value}'", db='Cheque.db')[0][0]
                filters.append("C.cheque_book_id = ?")
                cheque_params.append(CB_id)

            elif f_search_at == "Cheque No.":
                filters.append("C.Cheque_No LIKE '%' || ? || '%'")
                cheque_params.append(f_search_value)

            elif f_search_at == "Payee/Recipient":
                filters.append("C.Payee_name LIKE '%' || ? || '%'")
                cheque_params.append(f_search_value)
            
            elif f_search_at == "Cheque Type":
                if f_search_value == 'Account Pay':
                    filters.append("C.Cheque_Type = ?")
                    cheque_params.append("Account")
                elif f_search_value == 'Cross Pay':
                    filters.append("C.Cheque_Type = ?")
                    cheque_params.append("Cross")
                elif f_search_value == 'Select':
                    filters.append("C.Cheque_Type = ?")
                    cheque_params.append("False")
                else:
                    pass

        if not f_payment_type == "Select" and f_payment_type != "":
            # ["Select","Pending", "Deposited", "Bounced", "Overdue"]
            if f_payment_type == "Overdue":
                filters.append("DATE(SUBSTR(Cheque_Date, 7, 4) || '-' || SUBSTR(Cheque_Date, 4, 2) || '-' || SUBSTR(Cheque_Date, 1, 2)) < DATE('now', '-90 days') AND Statues = 'Pending'")
            else:
                filters.append("Statues = ?")
                cheque_params.append(f_payment_type)

        cheque_query = f"""
SELECT
    id, (SELECT book_name FROM ChequeBook CB WHERE CB.id = C.cheque_book_id ) AS Book_Name,
    Cheque_No, Cheque_Date, Payee_name, Amount, Cheque_Type, Statues
FROM Cheques C
        """
        where_clause = f" WHERE {' AND '.join(filters)}" if filters else ""

        query = cheque_query + where_clause
        params = cheque_params

        result = fetch_data(query, params, db='Cheque.db')

        if result:
            for row in tree.get_children():
                tree.delete(row)
            for row_data in result:
                tree.insert("", "end", values=list(row_data))
        
            treeview_data = result

        else:
            messagebox.showinfo("No Results", "No matching records found.", parent=window)
            reset_data()
            return []
        

    def export_to_excel():
        file_path = filedialog.asksaveasfilename(
            title="Export Invoices Data",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            parent=window
        )

        if file_path:
            try:
                global treeview_data
                data = treeview_data

                wb = Workbook()
                ws = wb.active
                ws.title = "Cheques List"

                headers = ['Id', 'Cheque Book Name', 'Cheque No.', 'Cheque Date', 'Payee Name', 'Amount', 'Cheque Type', 'Statues']
                ws.append(headers)

                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")

                for row in data:
                    ws.append(row)

                column_widths = {
                    'A': 10,  # Id
                    'B': 30,  # Cheque Book Name
                    'C': 20,  # Cheque No.
                    'D': 20,  # Cheque Date
                    'E': 35,  # Payee Name
                    'F': 18,  # Amount
                    'G': 18,  # Cheque Type
                    'H': 18,  # Statues
                }

                # Apply column widths
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width

                # Center-align all cells
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # Save the workbook to a file
                wb.save(file_path)

                # Show success message
                response = messagebox.askquestion("Data Exported", f"Excel saved at {file_path} \nDo you want to open the excel?", parent=window)
                if response == 'yes':
                    os.startfile(file_path)

            except PermissionError:
                # Handle the case when the file is open or locked
                messagebox.showerror("Error", "The file is currently open or locked. Please close it and try again.", parent=window)
            except Exception as e:
                # Catch any other unexpected errors
                messagebox.showerror("Error", f"An error occurred: {e}", parent=window)
        else:
            pass


    canvas = Canvas(
        window,
        bg = "#E7EBFF",
        height = 720,
        width = 1280,
        bd = 0,
        highlightthickness = 0,
        relief = "ridge"
    )
    canvas.place(x = 0, y = 0)

    start_date = DateEntry(
        window,
        bd=0,
        state=NORMAL,
        bg="#FFFFFF",
        fg="#000716",
        highlightthickness=0,
        date_pattern='dd/mm/yyyy',
    )
    start_date.place(
        x=60.0,
        y=56.0,
        width=120.0,
        height=25.0
    )

    start_date.bind("<FocusOut>", on_date_change)

    end_date = DateEntry(
        window,
        bd=0,
        state=NORMAL,
        bg="#FFFFFF",
        fg="#000716",
        highlightthickness=0,
        date_pattern='dd/mm/yyyy',
    )
    end_date.place(
        x=217.0,
        y=56.0,
        width=120.0,
        height=25.0
    )
    end_date.bind("<FocusOut>", on_date_change)

    date_options = [
    "All",
    "Today",
    "Last 7 Days",
    "Last 30 Days",
    "This Year",
    "Last Year",
    "Custom Range"
    ]

    quick_date = Combobox(
        window,
        values=date_options,
        state="readonly",  # Makes it read-only to prevent arbitrary input
        font=("VarelaRound Regular", 10)  # Styling the font
    )
    quick_date.set("Select")
    quick_date.place(
        x=374.0,
        y=56.0,
        width=150.0,
        height=25.0
    )
    quick_date.bind("<<ComboboxSelected>>", update_dates)


    search_at = Combobox(
        window,
        values=["Select", "Cheque Book", "Cheque No.", "Payee/Recipient", "Cheque Type"],
        state="readonly",
        font=("VarelaRound Regular", 10) 
    )
    search_at.set("Select")
    search_at.place(
        x=561.0,
        y=56.0,
        width=130.0,
        height=25.0
    )

    entry_image_5 = PhotoImage(
        file=relative_to_assets("entry_5.png"))
    entry_bg_5 = canvas.create_image(
        789.0,
        68.5,
        image=entry_image_5
    )
    
    search_at.bind("<<ComboboxSelected>>", update_search_field)
    update_search_field(None)


    payment_type = Combobox(
        window,
        values=["Select","Pending", "Deposited", "Bounced", "Overdue"],
        state="readonly",  
        font=("VarelaRound Regular", 10)  # Font styling
    )
    payment_type.set("Select")
    payment_type.place(
        x=911.0,
        y=56.0,
        width=170.0,
        height=25.0
    )


    button_image_1 = PhotoImage(
        file=relative_to_assets("button_1.png"))
    search_btn = Button(
        window,
        image=button_image_1,
        borderwidth=0,
        highlightthickness=0,
        command=filter_data,
        relief="flat"
    )
    search_btn.place(
        x=1121.0,
        y=51.0,
        width=30.0,
        height=30.0
    )

    button_image_2 = PhotoImage(
        file=relative_to_assets("button_2.png"))
    reset_btn = Button(
        window,
        image=button_image_2,
        borderwidth=0,
        highlightthickness=0,
        command=reset_data,
        relief="flat"
    )
    reset_btn.place(
        x=1190.0,
        y=51.0,
        width=30.0,
        height=30.0
    )


    canvas.create_text(
        60.0,
        37.0,
        anchor="nw",
        text="Starting Date",
        fill="#000000",
        font=("VarelaRound Regular", 12 * -1)
    )

    canvas.create_text(
        217.0,
        37.0,
        anchor="nw",
        text="Ending Date",
        fill="#000000",
        font=("VarelaRound Regular", 12 * -1)
    )

    canvas.create_text(
        191.0,
        60.0,
        anchor="nw",
        text="To",
        fill="#000000",
        font=("VarelaRound Regular", 13 * -1)
    )

    canvas.create_text(
        561.0,
        37.0,
        anchor="nw",
        text="Search as",
        fill="#000000",
        font=("VarelaRound Regular", 12 * -1)
    )

    canvas.create_text(
        704.0,
        37.0,
        anchor="nw",
        text="Search",
        fill="#000000",
        font=("VarelaRound Regular", 12 * -1)
    )

    canvas.create_text(
        374.0,
        37.0,
        anchor="nw",
        text="Quick Date",
        fill="#000000",
        font=("VarelaRound Regular", 12 * -1)
    )

    canvas.create_text(
        911.0,
        37.0,
        anchor="nw",
        text="Cheque status",
        fill="#000000",
        font=("VarelaRound Regular", 12 * -1)
    )

    button_image_4 = PhotoImage(
        file=relative_to_assets("button_5.png"))
    button_4 = Button(
        window,
        image=button_image_4,
        borderwidth=0,
        highlightthickness=0,
        command=PrintOnCheque,
        relief="flat"
    )
    button_4.place(
        x=960.0,
        y=106.0,
        width=120.0,
        height=30.0
    )

    button_image_3 = PhotoImage(
        file=relative_to_assets("button_3.png"))
    button_3 = Button(
        window,
        image=button_image_3,
        borderwidth=0,
        highlightthickness=0,
        command=export_to_excel,
        relief="flat"
    )
    button_3.place(
        x=1100.0,
        y=106.0,
        width=120.0,
        height=30.0
    )


    columns = ['Id', 'Cheque Book Name', 'Cheque No.', 'Cheque Date', 'Payee Name', 'Amount', 'Cheque Type', 'Statues']

    tree_frame = ttk.Frame(window)
    tree_frame.place(x=25, y=161, width=1230, height=540)

    x_scroll = ttk.Scrollbar(tree_frame, orient="horizontal")
    x_scroll.pack(side="bottom", fill="x")

    y_scroll = ttk.Scrollbar(tree_frame, orient="vertical")
    y_scroll.pack(side="right", fill="y")

    tree = ttk.Treeview(
        tree_frame,
        columns=columns,
        show="headings",
        xscrollcommand=x_scroll.set,
        yscrollcommand=y_scroll.set,
        height=25
    )

    x_scroll.config(command=tree.xview)
    y_scroll.config(command=tree.yview)

    for col in columns:
        tree.heading(col, text=col, anchor="center")
    
    columns_data = [
        ("Id", 50),
        ("Cheque Book Name", 230),
        ("Cheque No.", 140),
        ("Cheque Date", 150),
        ("Payee Name", 220),
        ("Amount", 150),
        ("Cheque Type", 130),
        ("Statues", 140)
    ]

    for name, width in columns_data:
        tree.column(name, anchor="center", width=width, stretch=False)


    style = ttk.Style(window)
    style.configure("Treeview.Heading", font=("VarelaRound Bold", 10), background="#D3D3D3")  
    style.configure("Treeview", rowheight=25,)
    style.map("Treeview", background=[("selected", "#3248ea")])

    tree.pack(expand=True, fill="both")


    def on_right_click(event):
        selected_item = tree.identify_row(event.y)
        if selected_item:
            tree.selection_set(selected_item)
            context_menu.post(event.x_root, event.y_root)


    def export_item():
        selected_item = tree.selection()
        if selected_item:
            row_data = tree.item(selected_item, "values")
            cheque_id = row_data[0]
            reset_data()
            PrintOnCheque(cheque_id, window)



    def delete_item():
        selected_item = tree.selection()
        if selected_item:
            row_data = tree.item(selected_item, "values")

            id = row_data[0] 

            confirm = messagebox.askyesno(
                "Confirm Delete",
                f"Are you sure you want to delete the Cheque with No.'{row_data[2]}'?",
                parent=window
            )
            if confirm:
                try:
                    run_query(f"DELETE FROM Cheques WHERE id = {id}", db='Cheque.db')
                    messagebox.showinfo("Success", "The Invoice is successfully Deleted!", parent=window)

                except sqlite3.Error as e:
                    messagebox.showerror(
                        "Database Error",
                        f"An error occurred while deleting the record: {e}"
                        , parent=window
                    )
                tree.delete(selected_item)

    def mark_as(text):
        selected_item = tree.selection()
        if selected_item:
            row_data = tree.item(selected_item, "values")

            id = row_data[0] 

            try:
                run_query(f"UPDATE Cheques SET Statues = '{text}' WHERE id = {id}", db='Cheque.db')

                messagebox.showinfo("Success", f"The Cheques is successfully {text}!", parent=window)

                tree.item(selected_item, values=(
                    row_data[0],  # Id
                    row_data[1],  # Cheque Book Name
                    row_data[2],  # Cheque No.
                    row_data[3],  # Cheque Date
                    row_data[4],  # Payee Name
                    row_data[5],  # Amount
                    row_data[6],  # Cheque Type
                    text,         # Statues
                ))

            except sqlite3.Error as e:
                messagebox.showerror(
                    "Database Error",
                    f"An error occurred while Updating the record: {e}",
                    parent=window
                )           

    context_menu = Menu(window, tearoff=0)
    context_menu.add_command(label="  Delete Cheque             ", command=delete_item)
    context_menu.add_command(label="  Reprint Cheque            ", command=export_item)
    context_menu.add_command(label="  Mark as Deposited         ", command=lambda: mark_as("Deposited"))
    context_menu.add_command(label="  Mark as Bounced           ", command=lambda: mark_as("Bounced"))

    # Bind right-click to Treeview
    tree.bind("<Button-3>", on_right_click)

    window.resizable(False, False)
    window.bind("<Escape>", on_close)
    window.protocol("WM_DELETE_WINDOW", on_close)
    window.bind("<Return>", lambda event: filter_data())
    window.bind('<Control-s>', lambda event: export_to_excel())
    window.bind('<Control-n>', lambda event: PrintOnCheque(parent=window))

    window.mainloop()

if __name__ == "__main__":
    Manage_Cheques_ui()    
