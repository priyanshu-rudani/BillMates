import sys
from pathlib import Path

if getattr(sys, 'frozen', False):  # Running as a PyInstaller EXE
    root_path = Path(sys.executable).parent
else:  # Running as a Python script
    root_path = Path(__file__).parent.parent
sys.path.append(str(root_path))
 
from utilities.path_utils import *
from datetime import datetime, timedelta
from tkinter import messagebox
import sqlite3
import os
from tkinter import messagebox, filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
from utilities.path_utils import get_dates_between

# ================================================
# Essential Function
# ================================================


def get_dates_before(client_id, date):
    try:
        end_date = datetime.strptime(date, "%d/%m/%Y")
        query = f"""
        SELECT DISTINCT date 
        FROM Invoices 
        WHERE client_id = {client_id}
        UNION 
        SELECT DISTINCT date 
        FROM Purchase 
        WHERE client_id = {client_id}
        UNION 
        SELECT DISTINCT date 
        FROM Payments
        WHERE entity_id = {client_id}
        """

        rows = fetch_data(query)
        valid_dates = []
        for row in rows:
            try:
                date_obj = datetime.strptime(row[0], "%d/%m/%Y")
                if date_obj < end_date:
                    valid_dates.append(date_obj)
            except ValueError:
                pass

        valid_dates.sort()
        sorted_dates = [date.strftime("%d/%m/%Y") for date in valid_dates]
        return sorted_dates if sorted_dates else []
    
    except Exception as e:
        messagebox.showerror("Unexpected Error", f"An unexpected error occurred: {e}")

def get_opening_balance(client_id, start_date):
    try:
        dates = get_dates_before(client_id, start_date)
        placeholders = ', '.join(['?'] * len(dates))
        client_name = get_client_name(client_id)
        client_type = get_client_type(client_name)

        if client_type not in ['Client', 'Supplier']:
            return 0.0
        else:
            pass

        ledger_query = f"""
            SELECT 
                Credited,
                Debited
            FROM (
                SELECT
                    0 AS Credited,
                    i.total AS Debited
                FROM Invoices i
                JOIN Clients c ON i.client_id = c.id
                WHERE i.client_id = {client_id} AND i.date IN ({placeholders}) AND c.client_type = 'Client'

                UNION ALL

                SELECT
                    0 AS Credited,
                    i.total AS Debited
                FROM Purchase i
                JOIN Clients c ON i.client_id = c.id
                WHERE i.client_id = {client_id} AND i.date IN ({placeholders}) AND c.client_type = 'Supplier'

                UNION ALL

                SELECT
                    COALESCE(
                        CASE 
                            WHEN i.entity_type = 'Income' AND c.client_type = 'Client' THEN i.total_amount
                            WHEN i.entity_type = 'Expense' AND c.client_type = 'Supplier' THEN i.total_amount
                            WHEN i.Expanse_type = 'Adjustment Credit' THEN i.total_amount
                        END, 0) AS Credited,
                    COALESCE(
                        CASE 
                            WHEN i.Expanse_type = 'Adjustment Debit' THEN i.total_amount
                        END, 0) AS Debited
                FROM Payments i
                JOIN Clients c ON i.entity_id = c.id
                WHERE i.entity_id = {client_id} AND i.date IN ({placeholders})
            ) AS combined_results;
        """

        customer_ledger = fetch_data(ledger_query, dates * 3)

        query = f"""
            SELECT
                opening_balance
            FROM
                Clients
            WHERE
                id = {client_id};
        """

        opening_balance = fetch_data(query)[0][0]
        
        if opening_balance is None or opening_balance == 0:
            balance = 0.0
        else:
            balance = float(opening_balance)

        for entry in customer_ledger:
            credit, debit = entry

            balance += credit
            balance -= debit

        return balance

    except sqlite3.Error as e:
        messagebox.showerror("Unexpected Error", f"Database error: {e}")
        return 0

    except Exception as e:
        messagebox.showerror("Unexpected Error", f"An unexpected error occurred: {e}")
        return 0


# ================================================================================================================
# Customer Spacific
# ================================================================================================================

# Client Sales Reports
def customer_sales(client_name, first_date, last_date):

    dates = get_dates_between(first_date, last_date)
    placeholders = ', '.join(['?'] * len(dates))
    case_statement = " ".join(
        [f"WHEN '{date}' THEN {index}" for index, date in enumerate(dates, start=1)]
    )
    client_id = get_client_id(client_name)

    query = f"""
        SELECT 
            i.invoice_no, 
            i.date, 
            c.name, 
            i.total, 
            i.paid, 
            i.remaining, 
            i.paid_status, 
            i.reference_no
        FROM 
            Invoices i 
        JOIN 
            Clients c ON i.client_id = c.id 
        WHERE 
            i.client_id = {client_id} AND i.date IN ({placeholders}) 
        ORDER BY 
            CASE i.date {case_statement} ELSE 999 END;
    """

    report = fetch_data(query, dates)

    query =f"""
    SELECT
        SUM(i.paid) AS Total_paid,
        SUM(i.remaining) AS Total_unpaid,
        SUM(i.total) AS Total
    FROM
        Invoices i
    JOIN 
        Clients c ON i.client_id = c.id
    WHERE 
        i.date IN ({placeholders}) AND i.client_id = {client_id}
    """
    total = fetch_data(query, dates)
    total = total[0]
    
    if total[0] is None:
        return []
    else:
        data = [
            ('', '', '', '', '', '', '', ''),
            ('', '', '', '', '', '', '', ''),
            ('Paid sales', total[0], '', 'Credit Sales', total[1], '', 'Total Sales', total[2]),
            ('', '', '', '', '', '', '', '')
        ]
        
        customer_sales_report = report + data
        return customer_sales_report

# Purchase Reports 
def purchase_report(client_name, first_date, last_date):
    dates = get_dates_between(first_date, last_date)
    placeholders = ', '.join(['?'] * len(dates))
    case_statement = " ".join(
        [f"WHEN '{date}' THEN {index}" for index, date in enumerate(dates, start=1)]
    )
    client_id = get_client_id(client_name)

    query = f"""
        SELECT 
            i.invoice_no, 
            i.date, 
            c.name, 
            i.total, 
            i.paid, 
            i.remaining, 
            i.paid_status, 
            i.reference_no
        FROM 
            Purchase i 
        JOIN 
            Clients c ON i.client_id = c.id 
        WHERE 
            i.client_id = {client_id} AND i.date IN ({placeholders}) 
        ORDER BY 
            CASE i.date {case_statement} ELSE 999 END;
    """

    report = fetch_data(query, dates)

    query =f"""
    SELECT
        SUM(i.paid) AS Total_paid,
        SUM(i.remaining) AS Total_unpaid,
        SUM(i.total) AS Total
    FROM
        Purchase i
    JOIN 
        Clients c ON i.client_id = c.id
    WHERE 
        i.date IN ({placeholders}) AND i.client_id = {client_id}
    """
    total = fetch_data(query, dates)
    total = total[0]

    if total[0] is None:
        return []
    else:
        data = [
            ('', '', '', '', '', '', '', ''),
            ('', '', '', '', '', '', '', ''),
            ('Paid sales', total[0], '', 'Credit Sales', total[1], '', 'Total Sales', total[2]),
            ('', '', '', '', '', '', '', '')
        ]

        Purchase_report = report + data
        return Purchase_report 

# Client Payment Reports
def customer_payment(client_name, first_date, last_date):
    dates = get_dates_between(first_date, last_date)
    placeholders = ', '.join(['?'] * len(dates))
    case_statement = " ".join(
        [f"WHEN '{date}' THEN {index}" for index, date in enumerate(dates, start=1)]
    )
    client_id = get_client_id(client_name)

    query = f"""
        SELECT 
            p.date AS Payment_Date,
            p.Expanse_type AS Payment_Type,
            c.name AS Client_Name, 
            p.total_amount AS Total_Amount, 
            p.payment_mode AS Payment_Mode,
            p.reference_no AS Reference_no
        FROM 
            Payments p
        JOIN 
            Clients c ON p.entity_id = c.id
        WHERE 
            p.entity_id = {client_id} AND p.date IN ({placeholders})
        ORDER BY 
            CASE p.date {case_statement} ELSE 999 END;
    """

    report = fetch_data(query, dates)

    query =f"""
    SELECT
        SUM(p.total_amount) AS Total_Amount
    FROM
        Payments P
    JOIN 
        Clients c ON P.entity_id = c.id
    WHERE 
        p.entity_id = {client_id} AND p.date IN ({placeholders})   
    """
    total = fetch_data(query, dates)
    total = total[0]

    if total[0] is None:
        return []
    else:
        data = [
            ('', '', '', '', '', ''),
            ('', '', '', '', '', ''),
            ('', '', 'Total Payment', total[0], '', ''),
            ('', '', '', '', '', '')
        ]

        customer_payment_report = report + data
        return customer_payment_report

# Client ledger
def customer_ledger(client_name, first_date, last_date):
    # Generate ledger query for the selected date range
    dates = get_dates_between(first_date, last_date)
    placeholders = ', '.join(['?'] * len(dates))
    case_statement = " ".join(
        [f"WHEN '{date}' THEN {index}" for index, date in enumerate(dates, start=1)]
    )
    client_id = get_client_id(client_name)
    client_type = get_client_type(client_name)

    if client_type not in ['Client', 'Supplier']:
        return []
    else:
        pass

    ledger_query = f"""
        SELECT 
            Date,
            Description,
            Credited,
            Debited
        FROM (
            SELECT 
                i.date AS Date,
                'Invoice (' || printf('INV%03d', i.invoice_no) || ')' AS Description,
                0 AS Credited,
                i.total AS Debited
            FROM Invoices i
            JOIN Clients c ON i.client_id = c.id
            WHERE i.client_id = {client_id} AND i.date IN ({placeholders}) AND c.client_type = 'Client'

            UNION ALL

            SELECT 
                i.date AS Date,
                'Invoice (' || printf('INV%03d', i.invoice_no) || ')' AS Description,
                0 AS Credited,
                i.total AS Debited
            FROM Purchase i
            JOIN Clients c ON i.client_id = c.id
            WHERE i.client_id = {client_id} AND i.date IN ({placeholders}) AND c.client_type = 'Supplier'

            UNION ALL

            SELECT 
                i.date AS Date,
                i.Expanse_type AS Description,
                COALESCE(
                    CASE 
                        WHEN i.entity_type = 'Income' AND c.client_type = 'Client' THEN i.total_amount
                        WHEN i.entity_type = 'Expense' AND c.client_type = 'Supplier' THEN i.total_amount
                        WHEN i.Expanse_type = 'Adjustment Credit' THEN i.total_amount
                    END, 0) AS Credited,
                COALESCE(
                    CASE 
                        WHEN i.Expanse_type = 'Adjustment Debit' THEN i.total_amount
                    END, 0) AS Debited
            FROM Payments i
            JOIN Clients c ON i.entity_id = c.id
            WHERE i.entity_id = {client_id} AND i.date IN ({placeholders})
        ) AS combined_results

        ORDER BY 
            CASE Date
                {case_statement}
                ELSE 999
            END;
    """

    customer_ledger = fetch_data(ledger_query, dates * 3)

    if not customer_ledger:
        return []

    balance = get_opening_balance(client_id, first_date)
    ledger_with_balance = []
    
    ledger_with_balance.append((' ', 'Opening Balance', 0, 0, balance))

    for entry in customer_ledger:
        date, description, credit, debit = entry

        balance += credit
        balance -= debit

        ledger_with_balance.append((date, description, credit, debit, balance))

    ledger_with_balance.append((' ', 'Closing Balance', 0, 0, balance))
    
    return ledger_with_balance

# cashflow Reports
def cashflow_report(first_date, last_date):
    dates = get_dates_between(first_date, last_date)
    placeholders = ', '.join(['?'] * len(dates))
    case_statement = " ".join(
        [f"WHEN '{date}' THEN {index}" for index, date in enumerate(dates, start=1)]
    )
    quary = f"""
        SELECT
            P.date AS Date,
            c.name AS Client_Name,
            CASE
                WHEN P.entity_type = 'Income' THEN 'Payment In'
                WHEN P.Expanse_type = 'Adjustment Credit' THEN 'Payment In'
                WHEN P.entity_type IN ('Expense', 'Salaries') THEN 'Payment Out'
                WHEN P.Expanse_type = 'Adjustment Debit' THEN 'Payment Out'
                ELSE 'Unknown'
            END AS Payment_Category,
            P.total_amount AS Total,
            P.payment_mode AS Payment_Mode,
            P.Expanse_type AS Payment_Type,
            P.reference_no AS Reference_No
        FROM 
            Payments P
        JOIN 
            Clients c ON P.entity_id = c.id
        WHERE 
            P.date IN ({placeholders})
        ORDER BY 
            CASE
                WHEN P.date IS NOT NULL THEN P.date
                {case_statement}
                ELSE 999
            END;
    """

    report = fetch_data(quary, dates)

    query =f"""
        SELECT 
            SUM(CASE 
                    WHEN entity_type  = 'Income' THEN total_amount
                    WHEN Expanse_type = 'Adjustment Credit' THEN total_amount
                END) AS payment_in,
            SUM(CASE 
                    WHEN entity_type  IN ('Expense', 'Salaries') THEN total_amount
                    WHEN Expanse_type = 'Adjustment Debit' THEN total_amount
                END) AS payment_out
        FROM Payments
        WHERE date IN ({placeholders})
    """
    total = fetch_data(query, dates)
    total = total[0]

    if total[0] is None:
        return []

    data = [
        ('', '', '', '', '', '', ''),
        ('', '', '', '', '', '', ''),
        ('', 'Total Payment In', total[0], '', 'Total Payment Out', total[1], ''),
        ('', '', '', '', '', '', ''),
    ]
        
    cashflow_report = report + data
    return cashflow_report

# profit and loss report
def profit_loss_report(first_date, last_date):
    dates = get_dates_between(first_date, last_date)
    placeholders = ', '.join(['?'] * len(dates))

    query = f"""
        WITH SalesData AS (
            SELECT 
                COALESCE(SUM(Paid), 0.0) AS collected_revenue,
                COALESCE(SUM(remaining), 0.0) AS accounts_receivable,
                COALESCE(SUM(total), 0.0) AS gross_revenue
            FROM Invoices
            WHERE date IN ({placeholders})
        ),
        PurchasesData AS (
            SELECT 
                COALESCE(SUM(Paid), 0.0) AS settled_purchases,
                COALESCE(SUM(remaining), 0.0) AS accounts_payable
            FROM Purchase
            WHERE date IN ({placeholders})
        ),
        SalariesData AS (
            SELECT 
                COALESCE(SUM(total_amount), 0.0) AS settled_salaries
            FROM Payments
            WHERE Expanse_type = 'Salary' AND date IN ({placeholders})
        ),
        LogisticsData AS (
            SELECT 
                COALESCE(SUM(total_amount), 0.0) AS logistics_costs
            FROM Payments
            WHERE Expanse_type = 'Transposition' AND date IN ({placeholders})
        ),
        ExpensesData AS (
            SELECT
                COALESCE(SUM(CASE WHEN Expanse_type = 'Rent' THEN total_amount ELSE 0 END), 0.0) AS rental_expenses,
                COALESCE(SUM(CASE WHEN Expanse_type = 'Utilities' THEN total_amount ELSE 0 END), 0.0) AS utilities_expenditure,
                COALESCE(SUM(CASE WHEN Expanse_type = 'Maintenance' THEN total_amount ELSE 0 END), 0.0) AS upkeep_expenses,
                COALESCE(SUM(CASE WHEN Expanse_type = 'Marketing' THEN total_amount ELSE 0 END), 0.0) AS Marketing_expenses,
                COALESCE(SUM(CASE WHEN Expanse_type = 'Loan Interest' THEN total_amount ELSE 0 END), 0.0) AS loan_interest_cost,
                COALESCE(SUM(CASE WHEN Expanse_type = 'Taxes' THEN total_amount ELSE 0 END), 0.0) AS statutory_dues,
                COALESCE(SUM(CASE WHEN Expanse_type = 'Other Expances' THEN total_amount ELSE 0 END), 0.0) AS other_expenses
            FROM Payments
            WHERE Expanse_type IN ('Rent', 'Utilities', 'Maintenance', 'Marketing', 'Loan Interest', 'Taxes', 'Other Expances') 
            AND entity_type = 'Expense' 
            AND date IN ({placeholders})
        )
        SELECT
            collected_revenue,
            accounts_receivable,
            gross_revenue,
            settled_purchases,
            accounts_payable,
            settled_salaries,
            logistics_costs,
            (
                settled_purchases + accounts_payable + settled_salaries + logistics_costs
            ) AS total_cogs,
            (
                gross_revenue - (settled_purchases + accounts_payable + settled_salaries + logistics_costs)
            ) AS Gross_Profit,
            rental_expenses,
            utilities_expenditure,
            upkeep_expenses,
            Marketing_expenses,
            loan_interest_cost,
            statutory_dues,
            other_expenses,
            (
                rental_expenses + utilities_expenditure + upkeep_expenses + Marketing_expenses + loan_interest_cost + statutory_dues + other_expenses
            ) AS toal_expenses,
            gross_revenue - (
                settled_purchases + accounts_payable + settled_salaries + logistics_costs +
                rental_expenses + utilities_expenditure + upkeep_expenses + Marketing_expenses + 
                loan_interest_cost + statutory_dues + other_expenses
            ) AS net_profit_loss
        FROM SalesData, PurchasesData, SalariesData, LogisticsData, ExpensesData
        """

    report = fetch_data(query, dates*5)
    report = report[0]

    if report[0] is None:
        return []
    
    data = [
        (' ', 'Revenue', '', ''),
        (1, 'Collected Revenue', '', round(report[0], 2)),
        (2, 'Accounts Receivable', '', round(report[1], 2)),
        (3, 'Gross Revenue', '(1+2)', round(report[2], 2)),
        (' ', '', '', ''),
        (' ', 'Cost of Goods solds', '', ''),
        (4, 'Settled Purchases', '', round(report[3], 2)),
        (5, 'Accounts Payable', '', round(report[4], 2)),
        (6, 'Settled Salaries', '', round(report[5], 2)),
        (7, 'Comprehensive Logistics Costs', '', round(report[6], 2)),
        (8, 'Total COGS Expenditure', '(4+5+6+7)', round(report[7], 2)),
        (' ', '', '', ''),
        (9, 'Gross Profit/Loss', '(3-8)', round(report[8], 2)),
        (' ', '', '', ''),
        (' ', 'Expenses', '', ''),
        (10, 'Rental Expenses', '', round(report[9], 2)),
        (11, 'Utilities Expenditure', '', round(report[10], 2)),
        (12, 'Upkeep Expenses', '', round(report[11], 2)),
        (13, 'Marketing Expenses', '', round(report[12], 2)),
        (14, 'Loan Interest Cost', '', round(report[13], 2)),
        (15, 'Statutory Dues', '', round(report[14], 2)), 
        (16, 'Other Expances', '', round(report[15], 2)),
        (17, 'Total Expenses', '(10 To 16)', round(report[16], 2)),
        (' ', '', '', ''),
        (18, 'Net Profit/Loss', '(9-17)', round(report[17], 2)),
    ]
    return data 

# Revenue Reports
def revenue_report(first_date, last_date):
    dates = get_dates_between(first_date, last_date)
    placeholders = ', '.join(['?'] * len(dates))
    case_statement = " ".join(
        [f"WHEN '{date}' THEN {index}" for index, date in enumerate(dates, start=1)]
    )
    query = f"""
    SELECT
        Payment_Date,
        type,
        client_name,
        Amount,
        payment_status,
        reference_no
    FROM(
        SELECT 
            i.date AS Payment_Date,
            'Sales Invoice(INV' || printf('%03d', i.invoice_no) || ')' AS type,
            c.name AS client_name,
            i.total AS Amount,
            i.paid_status AS payment_status,
            i.reference_no AS reference_no
        FROM 
            Invoices i
        JOIN 
            Clients c ON i.client_id = c.id
        WHERE 
            i.date IN ({placeholders})
    
        UNION ALL
    
        SELECT 
            p.date AS Payment_Date,
            p.Expanse_type AS type,
            c.name AS client_name,
            p.total_amount AS Amount,
            p.payment_mode AS payment_status,
            p.reference_no AS reference_no
        FROM 
            Payments p
        JOIN 
            Clients c ON p.entity_id = c.id
        WHERE 
            p.date IN ({placeholders}) AND p.Expanse_type = 'Adjustment Credit'
    ) AS subquery
    ORDER BY 
        CASE 
            Payment_Date
            {case_statement}
            ELSE 999
        END;
    """

    report = fetch_data(query, dates*2)

    query =f"""
    SELECT
        SUM(Total_Amount) AS Total_Amount
    FROM(
        SELECT 
            SUM(total_amount) AS Total_Amount 
        FROM Payments 
        WHERE 
            date IN ({placeholders}) AND Expanse_type = 'Adjustment Credit'

        UNION ALL

        SELECT 
            SUM(total) AS Total_Amount 
        FROM Invoices 
        WHERE 
            date IN ({placeholders})
    ) AS subquery
    """
    total = fetch_data(query, dates*2)

    if total[0] is None:
        return []
    
    data = [
        ('', '', '', '', '', ''),
        ('', '', '', '', '', ''),
        ('', '', '', 'Total Revenue', total[0][0], ''),
        ('', '', '', '', '', '')
    ]
        
    revenue_report = report + data
    return revenue_report

# Expense Analysis
def expense_analysis(first_date, last_date):
    dates = get_dates_between(first_date, last_date)
    placeholders = ', '.join(['?'] * len(dates))
    case_statement = " ".join(
        [f"WHEN '{date}' THEN {index}" for index, date in enumerate(dates, start=1)]
    )
    query = f"""
    SELECT
        Date,
        type,
        client_name,
        Amount,
        payment_mode,
        reference_no
    FROM(
        SELECT
            p.date AS Date,
            p.Expanse_type AS type,
            c.name AS client_name,
            p.total_amount AS Amount,
            p.payment_mode AS payment_mode,
            p.reference_no AS reference_no
        FROM 
            Payments p
        JOIN 
            Clients c ON p.entity_id = c.id
        WHERE
            p.date IN ({placeholders}) AND p.entity_type IN ('Expense', 'Salaries') OR Expanse_type = 'Adjustment Debit'
    ) AS subquery
    ORDER BY 
        CASE 
            Date
            {case_statement}
            ELSE 999
        END;

     """
    report = fetch_data(query, dates)

    query =f"""
    SELECT
        SUM(total_amount) AS Total_Amount
    FROM
        Payments
    WHERE
        date IN ({placeholders}) AND entity_type IN ('Expense', 'Salaries') OR Expanse_type = 'Adjustment Debit'
    """
    total = fetch_data(query, dates)

    if total[0] is None:
        return []

    data = [
        ('', '', '', '', '', ''),
        ('', '', '', '', '', ''),
        ('', '', 'Total Expanse', total[0][0], '', ''),
        ('', '', '', '', '', '')
    ]
        
    expense_analysis = report + data
    return expense_analysis



# Export as Excel
def export_report(report_type, start_date, end_date, client_name=None):
    try:
        safe_start_date = datetime.strptime(start_date, "%d/%m/%Y").strftime("%d-%m-%Y")
        safe_end_date = datetime.strptime(end_date, "%d/%m/%Y").strftime("%d-%m-%Y")
        if client_name is None:
            file_name = f"{report_type}-{safe_start_date}_{safe_end_date}"
        else:
            file_name = f"{client_name}-{report_type}_{safe_start_date}_{safe_end_date}"

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel files", "*.xlsx")],
            initialfile = file_name,
            title=f"Save {report_type}"
        )

        if file_path:

            wb = Workbook()
            sheet = wb.active
            sheet.title = report_type


            if report_type == 'Sales Reports':  
                headers  = ["Invoice No.", "Date", "Client Name", "Total", "Paid", "Remaining", "Paid Status", "Reference No."]
                data = customer_sales(client_name, start_date, end_date)
                column_widths = [17, 17, 17, 17, 17, 17, 17, 20]

            elif report_type == 'Purchase Reports':
                headers  = ["Invoice No.", "Date", "Suppiler Name", "Total", "Paid", "Unpaid", "Paid Status", "Reference No."]
                data = purchase_report(client_name, start_date, end_date)
                column_widths = [17, 17, 17, 17, 17, 17, 17, 20]

            elif report_type == 'Client Payment Reports':
                headers  = ["Date", "Payment Type", "Client Name", "Total Amount", "Payment Mode", "Reference No."]
                data = customer_payment(client_name, start_date, end_date)
                column_widths = [15, 20, 18, 17, 18, 21]

            elif report_type == 'Client Ledger':
                headers  = ["Date", "Description", "Credited", "Debited", "Balance"]
                data = customer_ledger(client_name, start_date, end_date)
                column_widths = [15, 35, 14, 14, 14]

            elif report_type == 'Cash Flow Statement':
                headers  = ["Date", "Client Name", "Payment Category", "Total", "Payment Mode", "Payment Type", "Reference No."]
                data = cashflow_report(start_date, end_date)
                column_widths = [15, 20, 20, 15, 20, 25, 20]

            elif report_type == 'Profit and Loss Statement':
                headers  = ["Sr No.", "Category", " ", "Amount"]
                data = profit_loss_report(start_date, end_date)
                column_widths = [10, 39, 10, 15]

            elif report_type == 'Revenue Reports':
                headers  = ["Payment Date", "Payment Type", "Client Name", "Amount", "Payment Status", "Reference no."]
                data = revenue_report(start_date, end_date)
                column_widths = [16, 25, 20, 16, 19, 20]

            elif report_type == 'Expense Analysis':
                headers  = ["Date", "Expense Type", "Client Name", "Amount", "Payment Mode", "Reference no."]
                data = expense_analysis(start_date, end_date)
                column_widths = [15, 30, 22, 16, 16, 20]

            else:
                headers = []
                data = []
                return


            for col_num, width in enumerate(column_widths, start=1):
                col_letter = get_column_letter(col_num)
                sheet.column_dimensions[col_letter].width = width

            sheet.row_dimensions[1].height = 22
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="000000", size= 12)
                cell.fill = PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row_num, row in enumerate(data, start=2):
                for col_num, value in enumerate(row, start=1):
                    cell = sheet.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            if report_type == 'Client Ledger':
                sheet.row_dimensions[2].height = 22
                for col_num in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=2, column=col_num)
                    cell.font = Font(bold=True, size= 12)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                last_row = sheet.max_row 
                sheet.row_dimensions[last_row].height = 22
                for col_num in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=last_row, column=col_num)
                    cell.font = Font(bold=True, size= 12)  
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for row in range(1, sheet.max_row + 1):  
                    cell = sheet.cell(row=row, column=2)  
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            elif report_type == 'Profit and Loss Statement':

                target_rows = [2, 5, 7, 12, 14, 16, 24, 26]
                for row_num in target_rows:
                    sheet.row_dimensions[row_num].height = 30
                    for col_num in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row_num, column=col_num)
                        cell.font = Font(bold=True) 

                for row_num in range(2, sheet.max_row + 1):
                    for col_num in range(1, sheet.max_column + 1):  # Loop through all columns
                        cell = sheet.cell(row=row_num, column=col_num)
                        if col_num == 2:  # Column B
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        elif col_num == 4:  # Column D
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.number_format = '#,##0.00'
                        elif col_num == 3:  # Column C
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:  # All other columns
                            cell.alignment = Alignment(horizontal="center", vertical="center")



            if report_type not in ['Client Ledger', 'Profit and Loss Statement']:
                total_rows = len(data) + 1
                if total_rows >= 3:
                    target_row = total_rows - 1
                    sheet.row_dimensions[target_row].height = 22
                    for col_num in range(1, len(headers) + 1):
                        cell = sheet.cell(row=target_row, column=col_num)
                        cell.font = Font(bold=True, color="000000", size= 12)
                        cell.fill = PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center") 

            wb.save(file_path)
            response = messagebox.askquestion("Invoice Saved", f"Invoice saved at {file_path} \nDo you want to open the Excel?")
            if response == 'yes':
                os.startfile(file_path)

    except PermissionError:
        messagebox.showerror("Permission Error", f"Permission denied: Unable to save the file. Please close the file")
    except FileNotFoundError:
        messagebox.showerror("File Not Found", f"Invalid file path: Please provide a valid path.")
    except Exception as e:
        messagebox.showerror("Unexpected Error", f"An unexpected error occurred while saving the file:\n{e}")