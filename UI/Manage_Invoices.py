import sys
from pathlib import Path

if getattr(sys, 'frozen', False):  # Running as a PyInstaller EXE
    root_path = Path(sys.executable).parent
else:  # Running as a Python script
    root_path = Path(__file__).parent.parent
sys.path.append(str(root_path))
 
from utilities.path_utils import *
import tkinter as tk
import sqlite3
import os
from tkinter import Canvas, Entry, ttk, Button, PhotoImage, messagebox, Menu, NORMAL, END, filedialog
from tkinter.ttk import Combobox
from tkcalendar import DateEntry
from utilities.inv import create_invoice
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from UI.gui import customer_payment_in, supplier_payments_out


def manage_invoice_ui(parent = None):

    ASSETS_PATH = Path(generate_path(root_path, "UI", "assets", "frame8"))
    
    def relative_to_assets(path: str) -> Path:
        return ASSETS_PATH / Path(path)

    window = tk.Toplevel(parent)
    window.focus()
    window.iconbitmap(generate_path("UI", "assets", "BillMates.ico"))
    window.transient(parent)
    window.grab_set()
    window.title("Manage Invoices")
    center_window(window, 1280, 720)
    window.configure(bg = "#E7EBFF")

    start_date = None
    end_date = None

    def on_date_change(event):
        quick_date.set("Custom Range")

    def update_dates(event):
        if start_date is None or end_date is None:
            return
        
        selected_option = quick_date.get()
        today = datetime.now()

        if selected_option == "Today":
            start_date.set_date(today)
            end_date.set_date(today)
            
        elif selected_option == "Last 7 Days":
            start_date.set_date(today - timedelta(days=7))
            end_date.set_date(today)
            
        elif selected_option == "Last 30 Days":
            start_date.set_date(today - timedelta(days=30))
            end_date.set_date(today)
            
        elif selected_option == "This Week":
            start_date.set_date(today - timedelta(days=today.weekday()))
            end_date.set_date(today)
            
        elif selected_option == "This Month":
            start_date.set_date(today.replace(day=1))
            end_date.set_date(today)
            
        elif selected_option == "This Year":
            start_date.set_date(today.replace(month=1, day=1))
            end_date.set_date(today)
            
        elif selected_option == "Last Year":
            start_date.set_date(today.replace(year=today.year - 1, month=1, day=1))
            end_date.set_date(today.replace(year=today.year - 1, month=12, day=31))
            
        elif selected_option == "Custom Range":
            # Enable the entry fields for manual date selection
            start_date.config(state=NORMAL)
            end_date.config(state=NORMAL)
            return   
        
    def update_search_field(event):
        global search_widget
        selected_option = search_at.get()

        # Destroy the previous widget (Entry or Combobox)
        for widget in window.winfo_children():
            if widget.winfo_name() == 'search_widget':
                widget.destroy()

        if selected_option == "Client Name":
            # Create a Combobox for Client Names
            client_names = [name[0] for name in fetch_data("SELECT name FROM Clients WHERE client_type != 'Staff'")]
            search_widget = ttk.Combobox(
                window,
                values=client_names,
                state="readonly",  # Read-only mode to prevent manual input
                font=("VarelaRound Regular", 10)
            )
            search_widget.place(
                x=704.0,
                y=56.0,
                width=170.0,
                height=25.0
            )
            search_widget.set("Select")

        elif selected_option == "Invoice Type":
            # Create a Combobox for Client Names
            client_names = ["Sales", "Purchase"]
            search_widget = ttk.Combobox(
                window,
                values=client_names,
                state="readonly",  # Read-only mode to prevent manual input
                font=("VarelaRound Regular", 10)
            )
            search_widget.place(
                x=704.0,
                y=56.0,
                width=170.0,
                height=25.0
            )
            search_widget.set("Select")
                          
        elif selected_option == "Payment Type":
            # Create a Combobox for Payment Types
            payment_types = ["Cash", "UPI", "Net Banking", "Check", "Card"]
            search_widget = ttk.Combobox(
                window,
                values=payment_types,
                state="readonly",  # Read-only mode to prevent manual input
                font=("VarelaRound Regular", 10)
            )
            search_widget.place(
                x=704.0,
                y=56.0,
                width=170.0,
                height=25.0
            )
            search_widget.set("Select")
        else:
            search_widget = Entry(
                window,
                bd=0,
                bg="#FFFFFF",
                fg="#000716",
                highlightthickness=0,
                font=("VarelaRound Regular", 10)
            )
            search_widget.place(
                x=704.0,
                y=56.0,
                width=170.0,
                height=25.0
            )

    def reset_data():
        global search_widget

        start_date.set_date(datetime.now().strftime("%d/%m/%Y"))
        end_date.set_date(datetime.now().strftime("%d/%m/%Y"))
        quick_date.set("Select")
        search_at.set("Select")
        search_widget.destroy()
        search_widget = Entry(
            window,
            bd=0,
            bg="#FFFFFF",
            fg="#000716",
            highlightthickness=0,
            font=("VarelaRound Regular", 10)
        )
        search_widget.place(
            x=704.0,
            y=56.0,
            width=170.0,
            height=25.0
        )
        payment_type.set("Select")
        
        for row in tree.get_children():
                    tree.delete(row)
 
    previous_data = {
        "start_date": datetime.now().strftime("%d/%m/%Y"),
        "end_date": datetime.now().strftime("%d/%m/%Y"),
        "quick_date": "Select",
        "search_at": "Select",
        "search_widget": "",
        "payment_type": "Select"
    }   

    def filter_data():
        global search_widget
        global treeview_data
        f_start_date = start_date.get()
        f_end_date = end_date.get()
        f_search_at = search_at.get()
        f_search_value = search_widget.get()
        f_payment_type = payment_type.get()

        try:
            common_columns = """
                invoice_no AS "Invoice No",
                (SELECT name FROM Clients cl WHERE cl.id = inv.client_id) AS "Client Name",
                date AS "Date",
                due_date AS "Due_date",
                total AS "Total",
                paid_status AS "Payment Status",
                payment_type AS "Payment Mode",
                paid AS "Paid",
                remaining AS "Remaining",
                reference_no AS "Reference No"
            """
            invoice_query = f"""
                SELECT {common_columns}
                FROM Invoices inv
            """
            purchase_query = f"""
                SELECT {common_columns}
                FROM Purchase inv
            """

            # Initialize filters and parameters
            filters = []
            invoice_params = []
            purchase_params = []

            if f_start_date != "" and f_end_date != "":
                start_d_obj = datetime.strptime(f_start_date, "%d/%m/%Y")
                end_d_obj = datetime.strptime(f_end_date, "%d/%m/%Y")

                # Generate a list of dates between start and end date (inclusive)
                date_list = []
                current_date = start_d_obj
                while current_date <= end_d_obj:
                    date_list.append(current_date.strftime("%d/%m/%Y"))  # Match the format in the database
                    current_date += timedelta(days=1)

                # Add filter for the 'date' column
                filters.append("date IN (" + ",".join(["?"] * len(date_list)) + ")")

                # Extend parameters with the generated date list
                invoice_params.extend(date_list)
                purchase_params.extend(date_list)

            if f_search_at != "Select" and f_search_at != "":
                if f_search_at == "Client Name":
                    client_id = get_client_id(f_search_value)
                    filters.append("inv.client_id = ?")
                    invoice_params.append(client_id)
                    purchase_params.append(client_id)
                elif f_search_at == "Invoice Type":
                    pass
                elif f_search_at == "Sales Inv No.":
                    filters.append("inv.invoice_no = ?")
                    invoice_params.append(f_search_value)
                elif f_search_at == "Purchase Inv No.":
                    filters.append("inv.invoice_no = ?")
                    purchase_params.append(f_search_value)
                elif f_search_at == "Payment Type":
                    filters.append("inv.payment_type = ?")
                    invoice_params.append(f_search_value)
                    purchase_params.append(f_search_value)
                elif f_search_value == "Select" or not f_search_value:
                    pass

            if f_payment_type != "Select" and f_payment_type != "":
                if f_payment_type == "Overdue":
                    filters.append("DATE(substr(due_date, 7, 4) || '-' || substr(due_date, 4, 2) || '-' || substr(due_date, 1, 2)) < DATE('now') AND remaining > 0")
                    
                else:
                    filters.append("paid_status = ?")
                    invoice_params.append(f_payment_type)
                    purchase_params.append(f_payment_type)

            where_clause = f" WHERE {' AND '.join(filters)}" if filters else ""

            if f_search_at == "Sales Inv No.":
                query = invoice_query + where_clause
                params = invoice_params
            elif f_search_at == "Purchase Inv No.":
                query = purchase_query + where_clause
                params = purchase_params
            elif f_search_at == "Invoice Type" and f_search_value != "Select":
                if f_search_value == "Sales":
                    query = invoice_query + where_clause
                    params = invoice_params
                elif f_search_value == "Purchase":
                    query = purchase_query + where_clause
                    params = purchase_params
            else:
                query = f"{invoice_query}{where_clause} UNION ALL {purchase_query}{where_clause}"
                params = invoice_params + purchase_params  # Combine params for both

            result = fetch_data(query, params)

            if result:
                for row in tree.get_children():
                    tree.delete(row)

                for idx, row_data in enumerate(result, start=1):
                    tree.insert("", "end", values=[idx] + list(row_data))

                previous_data["start_date"] = f_start_date
                previous_data["end_date"] = f_end_date
                previous_data["quick_date"] = quick_date.get()
                previous_data["search_at"] = f_search_at
                previous_data["search_widget"] = f_search_value
                previous_data["payment_type"] = f_payment_type
            
                treeview_data = result

            else:
                messagebox.showinfo("No Results", "No matching records found.", parent=window)

                start_date.delete(0, END)
                start_date.insert(0, previous_data["start_date"])
                end_date.delete(0, END)
                end_date.insert(0, previous_data["end_date"])
                quick_date.set(previous_data["quick_date"])
                search_at.set(previous_data["search_at"])
                payment_type.set(previous_data["payment_type"])
                return []
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"An error occurred: {e}", parent=window)
            return []

    def export_to_excel():
        file_path = filedialog.asksaveasfilename(
            title="Export Invoices Data",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            parent=window
        )

        if file_path:
            try:
                global treeview_data
                data = treeview_data

                wb = Workbook()
                ws = wb.active
                ws.title = "Invoices"

                headers = ['Invoice No.', 'Client Name', 'Invoice Date', 'Due Date', 'Amount', 'Status', 'Payment Mode', 'Discount', 'Total Amount', 'Remarks']
                ws.append(headers)

                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")

                for row in data:
                    formatted_row = ['INV' + str(row[0]).zfill(3)] + list(row[1:])
                    ws.append(formatted_row)

                column_widths = {
                    'A': 20,  # Invoice No
                    'B': 25,  # Client Name
                    'C': 18,  # Invoice Date
                    'D': 18,  # Due Date
                    'E': 20,  # Amount
                    'F': 15,  # Status
                    'G': 15,  # Payment Mode
                    'H': 15,  # Discount
                    'I': 20,  # Total Amount
                    'J': 30,  # Remarks
                }

                # Apply column widths
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width

                # Center-align all cells
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # Apply currency format to columns E, H, I
                currency_columns = ['E', 'H', 'I']
                for col_letter in currency_columns:
                    for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                        cell = ws[f"{col_letter}{row}"]
                        cell.number_format = '#,##,##0.00 ₹'

                # Save the workbook to a file
                wb.save(file_path)

                # Show success message
                response = messagebox.askquestion("Data Exported", f"Excel saved at {file_path} \nDo you want to open the excel?", parent=window)
                if response == 'yes':
                    os.startfile(file_path)

            except PermissionError:
                # Handle the case when the file is open or locked
                messagebox.showerror("Error", "The file is currently open or locked. Please close it and try again.", parent=window)
            except Exception as e:
                # Catch any other unexpected errors
                messagebox.showerror("Error", f"An error occurred: {e}", parent=window)
        else:
            pass

    def export_to_pdf():
        file_path = filedialog.asksaveasfilename(
            title="Export Invoices Data",
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf")],
            parent=window
        )
        if file_path:
            try:
                global treeview_data
                
                # Create PDF document
                doc = SimpleDocTemplate(file_path, pagesize=A4)
                elements = []

                headers = ['Invoice No.', 'Client Name', 'Invoice Date', 'Due Date', 'Amount', 'Status', 'Payment Mode', 'Discount', 'Total Amount']
                data = [headers] + [row[:-1] for row in treeview_data]
                # Create table with data
                table = Table(data)
                # Table styling
                table.setStyle(TableStyle([
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCC0DA')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                    ('TOPPADDING', (0, 0), (-1, 0), 5),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ]))
                # Add table to document
                elements.append(table)
                # Build PDF
                doc.build(elements)
                response = messagebox.askquestion("Data Exported", f"Excel saved at {file_path} \nDo you want to open the excel?", parent=window)
                if response == 'yes':
                    os.startfile(file_path)

            except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {e}", parent=window)

        else:
            pass


    canvas = Canvas(
        window,
        bg = "#E7EBFF",
        height = 720,
        width = 1280,
        bd = 0,
        highlightthickness = 0,
        relief = "ridge"
    )

    canvas.place(x = 0, y = 0)


    start_date = DateEntry(
        window,
        bd=0,
        state=NORMAL,
        bg="#FFFFFF",
        fg="#000716",
        highlightthickness=0,
        date_pattern='dd/mm/yyyy',
    )
    start_date.place(
        x=60.0,
        y=56.0,
        width=120.0,
        height=25.0
    )

    start_date.bind("<FocusOut>", on_date_change)



    end_date = DateEntry(
        window,
        bd=0,
        state=NORMAL,
        bg="#FFFFFF",
        fg="#000716",
        highlightthickness=0,
        date_pattern='dd/mm/yyyy',
    )
    end_date.place(
        x=217.0,
        y=56.0,
        width=120.0,
        height=25.0
    )
    end_date.bind("<FocusOut>", on_date_change)

    date_options = [
    "Today",
    "Last 7 Days",
    "Last 30 Days",
    "This Week",
    "This Month",
    "This Year",
    "Last Year",
    "Custom Range"
    ]

    quick_date = Combobox(
        window,
        values=date_options,
        state="readonly",  # Makes it read-only to prevent arbitrary input
        font=("VarelaRound Regular", 10)  # Styling the font
    )
    quick_date.set("Select")
    quick_date.place(
        x=374.0,
        y=56.0,
        width=150.0,
        height=25.0
    )
    quick_date.bind("<<ComboboxSelected>>", update_dates)


    search_at = Combobox(
        window,
        values=["Select","Client Name", "Invoice Type", "Sales Inv No.", "Purchase Inv No.", "Payment Type"],
        state="readonly",  # Makes it read-only to prevent arbitrary input
        font=("VarelaRound Regular", 10)  # Font styling to match the UI
    )
    search_at.set("Select")
    search_at.place(
        x=561.0,
        y=56.0,
        width=130.0,
        height=25.0
    )

    entry_image_5 = PhotoImage(
        file=relative_to_assets("entry_5.png"))
    entry_bg_5 = canvas.create_image(
        789.0,
        68.5,
        image=entry_image_5
    )

    
    search_at.bind("<<ComboboxSelected>>", update_search_field)
    update_search_field(None)


    payment_type = Combobox(
        window,
        values=["Select","Paid", "Unpaid", "Partially Paid", "Overdue"],  # Dropdown options
        state="readonly",  # Makes it read-only to prevent arbitrary input
        font=("VarelaRound Regular", 10)  # Font styling
    )
    payment_type.set("Select")
    payment_type.place(
        x=911.0,
        y=56.0,
        width=170.0,
        height=25.0
    )


    button_image_1 = PhotoImage(
        file=relative_to_assets("button_1.png"))
    search_btn = Button(
        window,
        image=button_image_1,
        borderwidth=0,
        highlightthickness=0,
        command=filter_data,
        relief="flat"
    )
    search_btn.place(
        x=1121.0,
        y=51.0,
        width=30.0,
        height=30.0
    )

    button_image_2 = PhotoImage(
        file=relative_to_assets("button_2.png"))
    reset_btn = Button(
        window,
        image=button_image_2,
        borderwidth=0,
        highlightthickness=0,
        command=reset_data,
        relief="flat"
    )
    reset_btn.place(
        x=1190.0,
        y=51.0,
        width=30.0,
        height=30.0
    )


    canvas.create_text(
        60.0,
        37.0,
        anchor="nw",
        text="Starting Date",
        fill="#000000",
        font=("VarelaRound Regular", 12 * -1)
    )

    canvas.create_text(
        217.0,
        37.0,
        anchor="nw",
        text="Ending Date",
        fill="#000000",
        font=("VarelaRound Regular", 12 * -1)
    )

    canvas.create_text(
        191.0,
        60.0,
        anchor="nw",
        text="To",
        fill="#000000",
        font=("VarelaRound Regular", 13 * -1)
    )

    canvas.create_text(
        561.0,
        37.0,
        anchor="nw",
        text="Search as",
        fill="#000000",
        font=("VarelaRound Regular", 12 * -1)
    )

    canvas.create_text(
        704.0,
        37.0,
        anchor="nw",
        text="Search",
        fill="#000000",
        font=("VarelaRound Regular", 12 * -1)
    )

    canvas.create_text(
        374.0,
        37.0,
        anchor="nw",
        text="Quick Date",
        fill="#000000",
        font=("VarelaRound Regular", 12 * -1)
    )

    canvas.create_text(
        911.0,
        37.0,
        anchor="nw",
        text="Payment status",
        fill="#000000",
        font=("VarelaRound Regular", 12 * -1)
    )



    button_image_3 = PhotoImage(
        file=relative_to_assets("button_3.png"))
    button_3 = Button(
        window,
        image=button_image_3,
        borderwidth=0,
        highlightthickness=0,
        command=export_to_excel,
        relief="flat"
    )
    button_3.place(
        x=960.0,
        y=106.0,
        width=120.0,
        height=30.0
    )

    button_image_4 = PhotoImage(
        file=relative_to_assets("button_4.png"))
    button_4 = Button(
        window,
        image=button_image_4,
        borderwidth=0,
        highlightthickness=0,
        command=export_to_pdf,
        relief="flat"
    )
    button_4.place(
        x=1100.0,
        y=106.0,
        width=120.0,
        height=30.0
    )

    

    columns = [
        "Sr no", "Invoice No", "Client Name", "Date", "Due_date", "Total", "Payment Status", "Payment Mode ", "Paid", "Remaining", "Reference No"
    ]

    tree_frame = ttk.Frame(window)
    tree_frame.place(x=25, y=161, width=1230, height=540)

    x_scroll = ttk.Scrollbar(tree_frame, orient="horizontal")
    x_scroll.pack(side="bottom", fill="x")

    y_scroll = ttk.Scrollbar(tree_frame, orient="vertical")
    y_scroll.pack(side="right", fill="y")

    tree = ttk.Treeview(
        tree_frame,
        columns=columns,
        show="headings",
        xscrollcommand=x_scroll.set,
        yscrollcommand=y_scroll.set,
        height=25
    )

    x_scroll.config(command=tree.xview)
    y_scroll.config(command=tree.yview)

    for col in columns:
        tree.heading(col, text=col, anchor="center")
    
    tree.column("Sr no", anchor="center", width=70, stretch=False)
    tree.column("Invoice No", anchor="center", width=150, stretch=False)
    tree.column("Client Name", anchor="center", width=250, stretch=False)
    tree.column("Date", anchor="center", width=190, stretch=False)
    tree.column("Due_date", anchor="center", width=190, stretch=False)
    tree.column("Total", anchor="center", width=130, stretch=False)
    tree.column("Payment Status", anchor="center", width=170, stretch=False)
    tree.column("Payment Mode ", anchor="center", width=130, stretch=False)
    tree.column("Paid", anchor="center", width=130, stretch=False)
    tree.column("Remaining", anchor="center", width=130, stretch=False)
    tree.column("Reference No", anchor="center", width=170, stretch=False)


    style = ttk.Style(window)
    style.configure("Treeview.Heading", font=("VarelaRound Bold", 10), background="#D3D3D3")  
    style.configure("Treeview", rowheight=25,)
    style.map("Treeview", background=[("selected", "#3248ea")])

    tree.pack(expand=True, fill="both")


    def on_right_click(event):
        selected_item = tree.identify_row(event.y)
        if selected_item:
            tree.selection_set(selected_item)
            context_menu.post(event.x_root, event.y_root)

    def export_item():
        selected_item = tree.selection()
        if selected_item:
            row_data = tree.item(selected_item, "values")

            invoice_no = row_data[1] 
            client_name = row_data[2]
            date = row_data[3]
            client_id = get_client_id(client_name)


            display_Inv_no = "INV" + str(invoice_no).zfill(3)

            inv_date = datetime.strptime(date, "%d/%m/%Y")
            safe_date = inv_date.strftime("%d-%m-%Y")

            filename = f"{display_Inv_no}_{safe_date}_{client_name}.pdf"
            file_path = filedialog.asksaveasfilename(
                title="Save PDF At",
                initialfile=filename,
                defaultextension=".pdf",  # Set the default extension
                filetypes=[("PDF Files", "*.pdf")],
                parent=window
            )

            if file_path:
                create_invoice(client_id, invoice_no, file_path)
                response = messagebox.askquestion("Invoice Saved", f"Invoice saved at {file_path} \nDo you want to open the PDF?", parent=window)
                if response == 'yes':
                    os.startfile(file_path)         

    def delete_item():
        selected_item = tree.selection()
        if selected_item:
            row_data = tree.item(selected_item, "values")

            invoice_no = row_data[1] 
            client_name = row_data[2]
            client_id = get_client_id(client_name)
            client_type = get_client_type(client_name)

            confirm = messagebox.askyesno(
                "Confirm Delete",
                f"Are you sure you want to delete the item with Invoice No '{invoice_no}' and Client Name '{client_name}'?",
                parent=window
            )
            if confirm:
                try:
                    if client_type == "Client":
                        run_query(
                            "DELETE FROM Invoices WHERE invoice_no = ? AND client_id = ?",
                            (invoice_no, client_id)  
                        )
                        run_query(
                            "DELETE FROM InvoiceItems WHERE invoice_no = ? AND client_id = ?",
                            (invoice_no, client_id)
                        )
                        messagebox.showinfo("Success", "The Invoice is successfully Deleted!", parent=window)
                    elif client_type == "Supplier":
                        run_query(
                            "DELETE FROM Purchase WHERE invoice_no = ? AND client_id = ?",
                            (invoice_no, client_id) 
                        )
                        run_query(
                            "DELETE FROM PurchaseItems WHERE invoice_no = ? AND client_id = ?",
                            (invoice_no, client_id)
                        )
                        messagebox.showinfo("Success", "The Invoice is successfully Deleted!", parent=window)

                except sqlite3.Error as e:
                    messagebox.showerror(
                        "Database Error",
                        f"An error occurred while deleting the record: {e}",
                        parent=window
                    )                
                tree.delete(selected_item)

    def mark_as_paid():
        selected_item = tree.selection()
        if selected_item:
            result = messagebox.askyesnocancel(
                        "Transcation Warning",
                        f"""You are about to mark this invoice as "Paid". 
                        \nThis action only changes the status label and will not automatically set the remaining balance to zero. To correctly clear the balance, you must record the full payment. 
                        \nHow would you like to proceed?
                        \n   YES to Record Full Payment
                        \n   NO to Change Status Only
                        \n   Cancel to Closes this window with no changes.
                        """,
                        parent=window
                     ) 
            
            if result == False:
                row_data = tree.item(selected_item, "values")

                invoice_no = row_data[1] 
                client_name = row_data[2]
                client_id = get_client_id(client_name)
                client_type = get_client_type(client_name)

                try:
                    if client_type == "Client":
                        run_query(
                            "UPDATE Invoices SET paid_status = 'Paid' WHERE invoice_no = ? AND client_id = ?",
                            (invoice_no, client_id)
                        )
                        messagebox.showinfo("Success", "The Invoice is successfully Mark as Paid!", parent=window)

                        tree.item(selected_item, values=(
                            row_data[0],  # Sr no
                            row_data[1],  # Invoice No
                            row_data[2],  # Client Name
                            row_data[3],  # Date
                            row_data[4],  # Due_date
                            row_data[5],  # Payment Mode
                            "Paid",        # Payment Status
                            row_data[7],  # Paid
                            row_data[8],  # Remaining
                            row_data[9]   # Reference No
                        ))

                    elif client_type == "Supplier":
                        run_query(
                            "UPDATE Purchases SET paid_status = 'Paid' WHERE invoice_no = ? AND client_id = ?",
                            (invoice_no, client_id)
                        )
                        messagebox.showinfo("Success", "The Invoice is successfully Mark as Paid!", parent=window)

                        tree.item(selected_item, values=(
                            row_data[0],  # Sr no
                            row_data[1],  # Invoice No
                            row_data[2],  # Client Name
                            row_data[3],  # Date
                            row_data[4],  # Due_date
                            row_data[5],  # Payment Mode
                            "Paid",        # Payment Status
                            row_data[7],  # Paid
                            row_data[8],  # Remaining
                            row_data[9]   # Reference No
                        ))

                except sqlite3.Error as e:
                    messagebox.showerror(
                        "Database Error",
                        f"An error occurred while deleting the record: {e}",
                        parent=window
                    )  

            elif result == True:
                selected_item = tree.selection()
                if selected_item:
                    row_data = tree.item(selected_item, "values")
                    client_name = row_data[2]
                    client_type = get_client_type(client_name)

                    if client_type == "Client":
                        customer_payment_in(window)
                    if client_type == "Supplier":
                        supplier_payments_out(window)
                    else:
                        return
                    

    def mark_as_Partially_paid():
        selected_item = tree.selection()
        if selected_item:
            result = messagebox.askyesnocancel(
                        "Transcation Warning",
                        f"""You are about to mark this invoice as "Partially Paid". 
                        \nThis action only changes the status label and will not automatically set the remaining balance to zero. To correctly clear the balance, you must record the payment. 
                        \nHow would you like to proceed?
                        \n   YES to Record Partial Payment
                        \n   NO to Change Status Only
                        \n   Cancel to Closes this window with no changes.
                        """,
                        parent=window
                     ) 
            
            if result == False:
                row_data = tree.item(selected_item, "values")

                invoice_no = row_data[1] 
                client_name = row_data[2]
                client_id = get_client_id(client_name)
                client_type = get_client_type(client_name)

                try:
                    if client_type == "Client":
                        run_query(
                            "UPDATE Invoices SET paid_status = 'Partially Paid' WHERE invoice_no = ? AND client_id = ?",
                            (invoice_no, client_id)
                        )
                        messagebox.showinfo("Success", "The Invoice is successfully Mark as Partially Paid!", parent=window)

                        tree.item(selected_item, values=(
                            row_data[0],  # Sr no
                            row_data[1],  # Invoice No
                            row_data[2],  # Client Name
                            row_data[3],  # Date
                            row_data[4],  # Due_date
                            row_data[5],  # Payment Mode
                            "Partially Paid",  # Payment Status
                            row_data[7],  # Paid
                            row_data[8],  # Remaining
                            row_data[9]   # Reference No
                        ))

                    elif client_type == "Supplier":
                        run_query(
                            "UPDATE Purchases SET paid_status = 'Partially Paid' WHERE invoice_no = ? AND client_id = ?",
                            (invoice_no, client_id)
                        )
                        messagebox.showinfo("Success", "The Invoice is successfully Mark as Partially Paid!", parent=window)

                        tree.item(selected_item, values=(
                            row_data[0],  # Sr no
                            row_data[1],  # Invoice No
                            row_data[2],  # Client Name
                            row_data[3],  # Date
                            row_data[4],  # Due_date
                            row_data[5],  # Payment Mode
                            "Partially Paid", # Payment Status
                            row_data[7],  # Paid
                            row_data[8],  # Remaining
                            row_data[9]   # Reference No
                        ))
                        
                except sqlite3.Error as e:
                    messagebox.showerror(
                        "Database Error",
                        f"An error occurred while deleting the record: {e}", 
                        parent=window
                    )
           
            elif result == True:
                selected_item = tree.selection()
                if selected_item:
                    row_data = tree.item(selected_item, "values")
                    client_name = row_data[2]
                    client_type = get_client_type(client_name)

                    if client_type == "Client":
                        customer_payment_in(window)
                    if client_type == "Supplier":
                        supplier_payments_out(window)
                    else:
                        return


    context_menu = Menu(window, tearoff=0)
    context_menu.add_command(label="  Export as PDF             ", command=export_item)
    context_menu.add_command(label="  Delete Invoice            ", command=delete_item)
    context_menu.add_command(label="  Mark as Paid              ", command=mark_as_paid)
    context_menu.add_command(label="  Mark as Partially Paid    ", command=mark_as_Partially_paid)

    # Bind right-click to Treeview
    tree.bind("<Button-3>", on_right_click)

    window.bind("<Return>", lambda event: filter_data())
    window.bind('<Control-s>', lambda event: export_to_excel())

    window.resizable(False, False)
    window.bind("<Escape>", lambda event: Top_Close(window, parent))
    window.protocol("WM_DELETE_WINDOW", lambda: Top_Close(window, parent))
    window.mainloop()

if __name__ == "__main__":
    manage_invoice_ui()    
